import openpyxl

def rows(filename):
	ws = openpyxl.load_workbook(filename, read_only=True).active
		
	for row_index in range(2, ws.max_row + 1):
		row = {}
		
		for col_index in range(1, ws.max_column + 1):
			key = str(ws.cell(row=1, column=col_index).value)
			val = str(ws.cell(row=row_index, column=col_index).value)
		
			row[key] = val

		yield row

def cols(filename):
	ws = openpyxl.load_workbook(filename, read_only=True).active
		
	for col_index in range(2, ws.max_column + 1):
		col = {}
		
		for row_index in range(1, ws.max_row + 1):
			key = str(ws.cell(row=row_index, column=1).value)
			val = str(ws.cell(row=row_index, column=col_index).value)
		
			col[key] = val

		yield col
